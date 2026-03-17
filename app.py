import os
import io
import re
import requests
import tempfile
import subprocess
from flask import Flask, request, jsonify, send_file, render_template
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl.utils

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-in-prod")

# ── Audio conversion (ffmpeg directly, no pydub) ──────────────────────────────
def webm_to_wav(webm_bytes):
    with tempfile.NamedTemporaryFile(suffix=".webm", delete=False) as inp:
        inp.write(webm_bytes)
        inp_path = inp.name
    out_path = inp_path.replace(".webm", ".wav")
    try:
        subprocess.run([
            "ffmpeg", "-y", "-i", inp_path,
            "-ac", "1", "-ar", "16000", "-sample_fmt", "s16",
            out_path
        ], check=True, capture_output=True)
        with open(out_path, "rb") as f:
            return f.read()
    finally:
        os.unlink(inp_path)
        if os.path.exists(out_path):
            os.unlink(out_path)

# ── Field parser ───────────────────────────────────────────────────────────────
def parse_fields(text):
    t = text.lower()
    tw = t.split()
    words = text.split()

    patterns = [
        (["referred by", "refer by", "referral", "referred"], "referred"),
        (["their mobile", "their number", "referrer mobile", "referrer number", "their mob"], "their_mobile"),
        (["referee mobile", "referee number", "referee mob", "refree mobile"], "ref_mobile"),
        (["referee", "refree", "referee name"], "referee"),
        (["mobile", "number"], "ref_mobile"),
    ]

    def find_kw(kws):
        for kw in sorted(kws, key=lambda x: -len(x.split())):
            kw_words = kw.split()
            n = len(kw_words)
            for i in range(len(tw) - n + 1):
                if tw[i:i+n] == kw_words:
                    return i, n
        return -1, 0

    anchors = {}
    for kws, field in patterns:
        if field not in anchors:
            idx, klen = find_kw(kws)
            if idx >= 0:
                anchors[field] = (idx, klen)

    sorted_anchors = sorted(anchors.items(), key=lambda x: x[1][0])
    result = {"referee": "", "ref_mobile": "", "referred": "", "their_mobile": ""}

    for i, (field, (idx, klen)) in enumerate(sorted_anchors):
        next_idx = sorted_anchors[i+1][1][0] if i+1 < len(sorted_anchors) else len(words)
        value_words = words[idx + klen: next_idx]
        value = " ".join(value_words).strip(",;.: ")
        if field in ("ref_mobile", "their_mobile"):
            value = re.sub(r"[^0-9+]", "", value.replace(" ", ""))
        result[field] = value

    return result

# ── Routes ─────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/transcribe", methods=["POST"])
def transcribe():
    api_key = request.headers.get("X-Sarvam-Key", "")
    if not api_key:
        return jsonify({"error": "Missing Sarvam API key"}), 400

    audio_file = request.files.get("audio")
    if not audio_file:
        return jsonify({"error": "No audio file"}), 400

    try:
        webm_bytes = audio_file.read()
        wav_bytes = webm_to_wav(webm_bytes)

        resp = requests.post(
            "https://api.sarvam.ai/speech-to-text",
            headers={"api-subscription-key": api_key},
            files={"file": ("audio.wav", io.BytesIO(wav_bytes), "audio/wav")},
            data={"model": "saarika:v2", "language_code": "en-IN"},
        )

        if not resp.ok:
            return jsonify({"error": f"Sarvam error {resp.status_code}: {resp.text}"}), 502

        transcript = resp.json().get("transcript", "")
        parsed = parse_fields(transcript)
        return jsonify({"transcript": transcript, "parsed": parsed})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/export", methods=["POST"])
def export():
    data = request.json
    entries = data.get("entries", [])
    if not entries:
        return jsonify({"error": "No entries"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    header_fill = PatternFill("solid", fgColor="1D9E75")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["#", "Referee Name", "Referee Mobile", "Referred By", "Their Mobile"]
    col_widths = [5, 24, 18, 24, 18]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    alt_fill = PatternFill("solid", fgColor="F0FAF6")

    for i, r in enumerate(entries, 1):
        row_data = [i, r.get("referee",""), r.get("ref_mobile",""), r.get("referred",""), r.get("their_mobile","")]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if i % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[i+1].height = 18

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name="tickets.xlsx")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
